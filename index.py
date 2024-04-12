import conexion
import exportarBD
import mostrar_about
import google_maps
from personalizarApariencia import PersonalizarApariencia

import tkinter as tk
from tkinter import filedialog, messagebox
import PyPDF2
import docx
import openpyxl
import json
import os
from odf.opendocument import load
from odf import text, teletype
from bs4 import BeautifulSoup
from SeleccionAparicion import SeleccionAparicion


personalizar_apariencia = None  # Variable para almacenar la instancia de PersonalizarApariencia

def abrir_ventana_personalizacion():
    global personalizar_apariencia
    if personalizar_apariencia is None:
        personalizar_apariencia = PersonalizarApariencia(root, text_box, text_box_porcion,lbl_inicio, frame_botones1, frame_botones2, btn_limpiar_busqueda, btn_seleccionar_porcion, btn_insertar_bd_desde_json, btn_ver_contenido_bd, entry_inicio)
    personalizar_apariencia.abrir_ventana_personalizacion()  # Llamar al método para abrir la ventana


def leer_archivo():
    global nombre_archivo
    nombre_archivo = filedialog.askopenfilename()
    try:
        if nombre_archivo:
            extension = os.path.splitext(nombre_archivo)[1].lower()
            with open(nombre_archivo, 'rb' if extension == '.pdf' else 'r') as archivo:
                contenido = ''
                if extension == '.pdf':
                    pdf_reader = PyPDF2.PdfReader(archivo)
                    for page_num in range(len(pdf_reader.pages)):
                        page = pdf_reader.pages[page_num]
                        contenido += page.extract_text()
                elif extension == '.docx':
                    doc = docx.Document(nombre_archivo)
                    for paragraph in doc.paragraphs:
                        contenido += paragraph.text
                elif extension == '.xlsx':
                    wb = openpyxl.load_workbook(nombre_archivo)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows(values_only=True):
                            contenido += ' '.join(str(cell) for cell in row) + '\n'
                elif extension == '.odt':
                    doc = load(nombre_archivo)
                    for para in doc.getElementsByType(text.P):
                        contenido += teletype.extractText(para) + '\n'
                elif extension == '.csv':
                    with open(nombre_archivo, 'r') as archivo_csv:
                        contenido = archivo_csv.read()
                elif extension == '.html':
                    with open(nombre_archivo, 'r') as archivo_html:
                        contenido = archivo_html.read()
                        soup = BeautifulSoup(contenido, 'html.parser')
                        contenido = soup.get_text()
                elif extension == '.json':
                    contenido = json.load(archivo)
                else:
                    contenido = archivo.read()

                text_box.delete('1.0', tk.END)
                text_box.insert(tk.END, contenido)
        else:
            print("No se seleccionó ningún archivo.")
    except FileNotFoundError:
        print("El archivo especificado no se encontró.")

def leer_archivo_odt(nombre_archivo):
    doc = load(nombre_archivo)
    contenido = ''
    for para in doc.getElementsByType(text.P):
        contenido += teletype.extractText(para) + '\n'
    return contenido

def leer_archivo_csv(nombre_archivo):
    with open(nombre_archivo, 'r') as archivo_csv:
        contenido = archivo_csv.read()
    return contenido

def leer_archivo_html(nombre_archivo):
    with open(nombre_archivo, 'r') as archivo_html:
        contenido = archivo_html.read()
        soup = BeautifulSoup(contenido, 'html.parser')
        return soup.get_text()






def seleccionar_porcion():
    global nombre_archivo
    contenido = text_box.get("1.0", "end-1c")
    inicio = entry_inicio.get()

    if not inicio:
        messagebox.showerror("Error", "Es necesario establecer un campo de búsqueda.")
        return

    if nombre_archivo:
        if inicio in contenido:
            indices = [i for i in range(len(contenido)) if contenido.startswith(inicio, i)]
            if indices:
                SeleccionAparicion(indices, contenido, inicio, text_box_porcion, text_box, nombre_archivo)
            else:
                messagebox.showinfo("Búsqueda", "El texto especificado no se encontró en el archivo.")
        else:
            messagebox.showinfo("Búsqueda", "El texto especificado no se encontró en el archivo.")
    else:
        print("Por favor, abre un archivo primero.")

def abrir_porcion():
    try:
        with open('porcion.json', 'r', encoding='utf-8') as f:  # Especifica la codificación UTF-8 al abrir el archivo
            contenido = json.load(f)
        mostrar_contenido(contenido)
    except FileNotFoundError:
        messagebox.showinfo("Archivo no encontrado", "El archivo JSON no existe.")

def mostrar_contenido(contenido):
    ventana = tk.Toplevel()
    ventana.title("Contenido del archivo JSON")

    text_box = tk.Text(ventana)
    text_box.pack(fill=tk.BOTH, expand=True)

    # Convertir el contenido JSON a una cadena y mostrarlo en el cuadro de texto
    contenido_str = json.dumps(contenido, indent=4, ensure_ascii=False)  # Asegúrate de no escapar los caracteres especiales
    text_box.insert(tk.END, contenido_str)
       

def limpiar_busqueda():
    entry_inicio.delete(0, tk.END)
    text_box_porcion.delete('1.0', tk.END)
    text_box.delete('1.0', tk.END)

def insertar_en_bd_desde_json():
    global nombre_archivo
    contenido = text_box_porcion.get("1.0", "end-1c")

    if not contenido:
        messagebox.showerror("Error", "Primero se debe seleccionar un texto desde un archivo.")
        return

    conexion.insertar_en_bd_desde_json()


ventana_contenido_bd = None

# Función para ver el contenido de la base de datos
def ver_contenido_bd():
    global ventana_contenido_bd

    try:
        registros = conexion.obtener_contenido_bd()
        
        def exportar_html():
            try:
                registros = conexion.obtener_contenido_bd()  # Obtener los registros aquí
                nombre_archivo_html = filedialog.asksaveasfilename(defaultextension=".html", filetypes=[("HTML Files", "*.html")])
                if nombre_archivo_html:
                    exportarBD.exportar_contenido_html(registros, nombre_archivo_html)
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al exportar a HTML: {str(e)}")
            
        
        if registros:
            if ventana_contenido_bd:
                ventana_contenido_bd.destroy()

            ventana_contenido_bd = tk.Toplevel(root)
            ventana_contenido_bd.title("Contenido de la Base de Datos")
            ventana_contenido_bd.geometry("500x500")

            menu_exportar = tk.Menu(ventana_contenido_bd)
            ventana_contenido_bd.config(menu=menu_exportar)

            submenu_exportar = tk.Menu(menu_exportar, tearoff=0)
            menu_exportar.add_cascade(label="Archivo", menu=submenu_exportar)
            submenu_exportar.add_command(label="Exportar a CSV", command=exportarBD.exportar_csv)
            submenu_exportar.add_command(label="Exportar a PDF", command=exportarBD.exportar_pdf)
            submenu_exportar.add_command(label="Exportar a HTML", command=exportarBD.exportar_contenido_html)
            
            submenu_exportar.add_separator()
            submenu_exportar.add_command(label="Cerrar Ventana", command=ventana_contenido_bd.destroy)

            frame_widgets = tk.Frame(ventana_contenido_bd)
            frame_widgets.pack(fill=tk.BOTH, expand=True)

            scrollbar_y = tk.Scrollbar(frame_widgets, orient="vertical")
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

            inner_frame = tk.Frame(frame_widgets)
            inner_frame.pack(fill=tk.BOTH, expand=True)

            text_box_frame = tk.Text(inner_frame, wrap="word", yscrollcommand=scrollbar_y.set, state="normal")
            text_box_frame.pack(fill="both", expand=True)

            scrollbar_y.config(command=text_box_frame.yview)

            
            
            for registro in registros:
                archivo, contenido, notas, fecha_guardado = registro
                
                # Obtener la fecha en formato español
                fecha_guardado_espanol = fecha_guardado.strftime('%d-%m-%Y %H:%M:%S')
                
                # Insertar el nombre del archivo en negrita
                text_box_frame.insert(tk.END, f"Archivo: {archivo}\n", "bold")
                text_box_frame.tag_configure("bold", font=("Arial", 10, "bold"))
                
                # Insertar fecha de guardado en negrita
                text_box_frame.insert(tk.END, f"Fecha de guardado: {fecha_guardado_espanol}\n", "bold")
                text_box_frame.tag_configure("bold", font=("Arial", 10, "bold"))

                # Insertar el contenido y las notas
                text_box_frame.insert(tk.END, contenido + "\n")

                if notas is not None:
                    text_box_frame.insert(tk.END, f"Nota: {notas}\n\n")

                # Insertar botón para ver en Google Maps
                if "Coordenadas:" in contenido or "coordenadas:" in contenido:
                    latitud, longitud = google_maps.obtener_coordenadas(contenido)
                    if latitud and longitud:
                        btn_abrir_mapa = tk.Button(inner_frame, text="Ver en Google Maps",
                                                   command=lambda lat=latitud, lon=longitud: google_maps.abrir_google_maps(lat, lon))
                        text_box_frame.window_create(tk.END, window=btn_abrir_mapa)

                        # Cambiar el cursor a flecha al pasar por encima de los botones
                        btn_abrir_mapa.bind("<Enter>", lambda event: btn_abrir_mapa.config(cursor="arrow"))
                        btn_abrir_mapa.bind("<Leave>", lambda event: btn_abrir_mapa.config(cursor=""))

                # Insertar botón para ver/añadir nota
                btn_nota = tk.Button(inner_frame, text="Ver/Añadir Nota", command=lambda arch=archivo: abrir_ventana_nota(arch))
                text_box_frame.window_create(tk.END, window=btn_nota)

                
                # Botón para eliminar registro                
                btn_eliminar = tk.Button(inner_frame, text="Eliminar Registro", command=lambda arch=archivo: eliminar_registro(arch))
                text_box_frame.window_create(tk.END, window=btn_eliminar)
                text_box_frame.insert(tk.END, "\n")

                # Insertar línea continua como separador
                text_box_frame.insert(tk.END, "-"*50 + "\n\n")

                # Cambiar el cursor a flecha al pasar por encima de los botones
                btn_nota.bind("<Enter>", lambda event: btn_nota.config(cursor="arrow"))
                btn_nota.bind("<Leave>", lambda event: btn_nota.config(cursor=""))

        else:
            messagebox.showinfo("Información", "No se encontraron registros en la base de datos.")
            if ventana_contenido_bd:
                ventana_contenido_bd.destroy()
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al obtener el contenido de la base de datos: {str(e)}")

# Función para eliminar un registro
def eliminar_registro(archivo):
    conexion.eliminar_registro_bd(archivo, ver_contenido_bd)



def abrir_ventana_nota(archivo):
    def ajustar_tamano(event):
        text_nota.config(width=ventana_nota.winfo_width() // 10, height=ventana_nota.winfo_height() // 25)

    def guardar_nota():
        nota = text_nota.get("1.0", tk.END).strip()
        if nota != "":
            conexion.guardar_nota_en_bd(archivo, nota)
            ventana_nota.destroy()
            if hasattr(root, 'ventana_contenido_bd') and tk.Toplevel in root.ventana_contenido_bd.__class__.__mro__:
                root.ventana_contenido_bd.destroy()
            ver_contenido_bd()
        else:
            if messagebox.askokcancel("Advertencia", "¿Estás seguro de guardar una nota vacía?"):
                conexion.guardar_nota_en_bd(archivo, nota)
                ventana_nota.destroy()
                if hasattr(root, 'ventana_contenido_bd') and tk.Toplevel in root.ventana_contenido_bd.__class__.__mro__:
                    root.ventana_contenido_bd.destroy()
                ver_contenido_bd()

    nota_existente = conexion.obtener_nota_desde_bd(archivo)

    ventana_nota = tk.Toplevel(root)
    ventana_nota.title(f"Nota para {archivo}")
    ventana_nota.geometry("400x300")  # Establecer tamaño específico

    label_nota = tk.Label(ventana_nota, text="Nota:")
    label_nota.pack()

    text_nota = tk.Text(ventana_nota, wrap="word")
    text_nota.pack(fill="both", expand=True)

    ventana_nota.bind("<Configure>", ajustar_tamano)

    if nota_existente:
        text_nota.insert(tk.END, nota_existente)

    btn_guardar_nota = tk.Button(ventana_nota, text="Guardar Nota", command=guardar_nota)
    btn_guardar_nota.pack(pady=5)


def eliminar_json():
    try:
        if os.path.exists('porcion.json'):
            confirmacion = messagebox.askyesno("Confirmar eliminación", "¿Estás seguro de que deseas eliminar el archivo porcion.json?")
            if confirmacion:
                os.remove('porcion.json')
                messagebox.showinfo("Información", "El archivo porcion.json se eliminó correctamente.")
        else:
            messagebox.showinfo("Información", "El archivo porcion.json no existe.")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al intentar eliminar el archivo porcion.json: {str(e)}")

root = tk.Tk()
root.title("Seleccionar información de un Archivo")

menu_principal = tk.Menu(root)
root.config(menu=menu_principal,  borderwidth=0)

menu_archivo = tk.Menu(menu_principal, tearoff=False, border=0)
menu_principal.add_cascade(label="Archivo", menu=menu_archivo)
menu_archivo.add_command(label="Abrir Archivo", command=leer_archivo)
menu_archivo.add_command(label="Personalizar Apariencia", command=abrir_ventana_personalizacion)
menu_archivo.add_separator()
menu_archivo.add_command(label="Eliminar archivo JSON", command=eliminar_json)
menu_archivo.add_separator()
menu_archivo.add_command(label="Salir", command=root.quit)

menu_about = tk.Menu(menu_principal, tearoff=False)
menu_principal.add_cascade(label="About", menu=menu_about)
menu_about.add_command(label="Acerca de", command=mostrar_about.mostrar_about)

text_box = tk.Text(root, height=20, width=50) 
text_box.pack(pady=5, fill=tk.BOTH, expand=True)

# Entrada de texto y etiqueta
lbl_inicio = tk.Label(root, text="Búsqueda:")
lbl_inicio.pack()
entry_inicio = tk.Entry(root, width=30)
entry_inicio.pack()

# Crear un frame para los botones de la primera fila
frame_botones2 = tk.Frame(root)
frame_botones2.pack(pady=5)  # Alineado justo debajo del entry_inicio

# Botones de la primera fila dentro del Frame
btn_limpiar_busqueda = tk.Button(frame_botones2, text="Limpiar", command=limpiar_busqueda)
btn_limpiar_busqueda.grid(row=0, column=0, padx=5)

btn_seleccionar_porcion = tk.Button(frame_botones2, text="BUSCAR Y GUARDAR en JSON", command=seleccionar_porcion)
btn_seleccionar_porcion.grid(row=0, column=1, padx=5)

btn_abrir_porcion = tk.Button(frame_botones2, text="Abrir Archivo JSON", command=abrir_porcion)
btn_abrir_porcion.grid(row=0, column=2, padx=5)


# Caja de texto en una fila separada
text_box_porcion = tk.Text(root, height=5, width=75)
text_box_porcion.pack(pady=5)

# Crear un frame para los botones de la primera fila
frame_botones1 = tk.Frame(root)
frame_botones1.pack(side=tk.BOTTOM, pady=5)  # Alineado en la parte baja de la ventana

# Botones de la segunda fila dentro del Frame
btn_insertar_bd_desde_json = tk.Button(frame_botones1, text="Insertar en Base de Datos", command=insertar_en_bd_desde_json)
btn_insertar_bd_desde_json.pack(side=tk.LEFT, padx=5, pady=5)

btn_ver_contenido_bd = tk.Button(frame_botones1, text="Ver Contenido de la Base de Datos", command=ver_contenido_bd)
btn_ver_contenido_bd.pack(side=tk.LEFT, padx=5, pady=5)

root.mainloop()

